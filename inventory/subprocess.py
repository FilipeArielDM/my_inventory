
import subprocess
import openpyxl
import os
from openpyxl import load_workbook, Workbook
import sys

# Versão onde o teste precisa ser feito em máquinas que não possui antivírus para verificar o comportamento para que caso haja erro,
# refazer a condição para que seja acrescentado None

# Caminho para o arquivo Excel no flash drive (ajuste conforme necessário)
excel_file_path = 'F:/informacoes_do_sistema.xlsx' # Supondo F: como a unidade do flash drive

def get_ipconfig_data():
    try:
        # Executar o comando ipconfig
        output = subprocess.check_output(['ipconfig'], text=True, shell=True)
        return output
    except subprocess.CalledProcessError as e:
        return None
    
def parse_ipconfig_data(ipconfig_output):
    lines = ipconfig_output.split('\n')
    ipv4_address = None
    for line in lines:
        if "IPv4" in line:  # Este texto pode precisar ser ajustado dependendo do idioma do sistema
            ipv4_address = line.split(': ')[1]
            break  # Remove this break if you want to collect IPs of all adapters
    return ipv4_address

ipconfig_data = get_ipconfig_data()
ipv4_address = parse_ipconfig_data(ipconfig_data) if ipconfig_data else None

# Funções execute_command e get_filtered_software como antes
def add_data_to_excel(file_path, data):
    # Verifica se o arquivo já existe
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        # Cria um novo arquivo e adiciona cabeçalhos
        wb = Workbook()
        ws = wb.active
        ws.append(['Host Name', 'OS Name', 'System Model', 'System Manufacturer', 'Softwares', 'IP'])
    
    # Determinar a próxima linha disponível para escrita
    new_row = ws.max_row + (1 if ws.max_row > 1 or ws['A1'].value else 0)
    
    # Adiciona nova linha de dados
    for index, value in enumerate(data, start=1):
        ws.cell(row=new_row, column=index, value=value)

    # Salvar as alterações
    wb.save(file_path)

# Função para executar comandos no shell e retornar a saída
def execute_command(command):
    try:
        result = subprocess.run(command, text=True, capture_output=True, shell=True)
        return result.stdout
    except subprocess.CalledProcessError as e:
        return str(e.output)

# Função para analisar a saída do systeminfo a buscar por informações específicas
def parse_systeminfo(info_list, data):
    for line in data.split('\n'):
        for info in info_list:
            if line.startswith(info):
                yield line

def get_filtered_software():
    try:
        # Executa o comando e obtém o resultado
        # Note: universal_newlines=True é equivalente a text=True em versões mais novas do Python (3.7+)
        # Isso faz com que a saída do comando seja tratada como texto
        result = subprocess.check_output('wmic product get name', shell=True, universal_newlines=True)

        # Filtra os nomes dos softwares
        filtered_names = [name.strip() for name in result.splitlines() 
                          if "TXOne" in name or "Symantec" in name or "CrowdStrike" in name]
        
        # Retorna os nomes filtrados, unidos por vírgula
        return ', '.join(filtered_names)
    
    except subprocess.CalledProcessError as e:
        # Retorna uma mensagem de erro caso o comando externo falhe
        return f"Erro ao executar comando: {e}"

# Chamada da função
filtered_software = get_filtered_software()

# Seu código para executar os comandos e coletar as saídas (como `execute_command`)
systeminfo_data = execute_command('systeminfo')

# Sua função `parse_systeminfo` para extrair linhas específicas da saída do comando
info_list = ["Host Name", "OS Name", "System Model", "System Manufacturer",'IP']
parsed_data_lines = list(parse_systeminfo(info_list, systeminfo_data))

# Executa systeminfo e filtra a saída
systeminfo_data = execute_command('systeminfo')
info_list = ["Host Name","OS Name", "System Model", "System Manufacturer", 'Software', 'IP']
parsed_data = list(parse_systeminfo(info_list, systeminfo_data))

# Cria uma nova planilha do Excel
wb = openpyxl.Workbook()
ws = wb.active
new_row = ws.max_row + 1

# Insere informações no Excel
ws['A1'] = 'Host Name'
ws['A2'] = parsed_data[0].split(':', 1)[1].strip() if len(parsed_data) > 0 else ''

ws['B1'] = 'OS Name'
ws['B2'] = parsed_data[1].split(':', 1)[1].strip() if len(parsed_data) > 1 else ''

ws['C1'] = 'System Model'
ws['C2'] = parsed_data[2].split(':', 1)[1].strip() if len(parsed_data) > 2 else ''

ws['D1'] = 'System Manufacturer'
ws['D2'] = parsed_data[3].split(':', 1)[1].strip() if len(parsed_data) > 3 else ''

ws['E1'] = 'Softwares'
ws['E2'] = parsed_data[4].split(':', 1)[1].strip() if len(parsed_data) > 4 else filtered_software

ws['F1'] = 'IP'
ws['F2'] = parsed_data[5].split(':', 1)[1].strip() if len(parsed_data) > 5 else ipv4_address

data_to_add = [
    parsed_data_lines[0].split(':', 1)[1].strip() if len(parsed_data_lines) > 0 else '',  # Host Name
    parsed_data_lines[1].split(':', 1)[1].strip() if len(parsed_data_lines) > 1 else '',  # OS Name
    parsed_data_lines[2].split(':', 1)[1].strip() if len(parsed_data_lines) > 2 else '',  # System Model
    parsed_data_lines[3].split(':', 1)[1].strip() if len(parsed_data_lines) > 3 else '',  # System Manufacturer
    filtered_software, #Software
    ipv4_address]  # IP

# Ajusta a largura das colunas
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 75
ws.column_dimensions['F'].width = 30

# Salva a planilha
wb.save('informacoes_do_sistema.xlsx')

# Chama a função para adicionar os dados coletados à planilha no caminho especificado
add_data_to_excel(excel_file_path, data_to_add)
